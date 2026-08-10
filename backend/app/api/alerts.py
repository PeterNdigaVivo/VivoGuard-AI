"""/alerts endpoints — list with filters, server-rendered titles, snapshot.

May-2026 redesign: every alert returned by /alerts includes
`title`, `body`, `severity`, and `snapshot_url` pre-computed by the
server. The frontend stops translating detection_type strings (which
made the chain /alerts page and the per-store feed render
inconsistently) and just renders what we send.
"""
from __future__ import annotations
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import JSONResponse

from app.utils.cache import cached_store_endpoint
from sqlalchemy import case, desc, func
from sqlalchemy.orm import Session

from app.config import settings
from app.database import get_db
from app.deps import get_current_user, require_role
from app.models import Alert, Camera, DetectionEvent, User, Zone
from app.schemas.alert import AlertActionOut, AlertNoteIn, AlertOut

router = APIRouter(prefix="/alerts", tags=["alerts"])


# ---- Title / body / severity translation ---------------------------
#
# Single source of truth for the human-readable alert presentation.
# Same logic feeds the per-store dashboard feed AND the chain /alerts
# page so labels are consistent.

# Severity bucket per detection_type. Front-end pulls from this for
# the colour (red / amber / blue dots).
_SEVERITY: dict[str, str] = {
    # critical — security / safety incidents
    "fight":             "critical",
    "intrusion":         "critical",
    "weapon":            "critical",
    "weapon_brandished": "critical",
    "fall":              "critical",
    "fire":              "critical",
    "smoke":             "critical",
    "shrinkage":         "critical",
    # warning — operational risks
    "queue":             "warning",
    "queue_length":      "warning",
    "crowd":             "warning",
    "trespass":          "warning",
    "loitering":         "warning",
    "shutter":           "warning",
    "abandoned_object":  "warning",
    "tailgating":        "warning",
    # info — routine operational signals
    "staff_present":     "info",
    "occupancy":         "info",
    "entry_exit":        "info",
    "dwell":             "info",
    "passersby":         "info",
    "live_activity":     "warning",
    "shop_open_close":   "info",
    "sales_floor_insight": "info",
    "store_intelligence":  "info",
}

# Feed-ordering rank lists are derived from the 4-tier _SEVERITY_4 ladder
# (defined below) so ALL high-priority types surface — see _RANK_CRITICAL /
# _RANK_HIGH after the _SEVERITY_4 definition.


def _severity(detection_type: str | None) -> str:
    return _SEVERITY.get(detection_type or "", "info")


# Non-technical traffic-light labels. Spec Part 3 mapping:
#   URGENT (red)    — act now
#   ATTENTION (amber) — act within ~15 min
#   INFO (blue)     — for the record
_SEVERITY_LABEL: dict[str, str] = {
    "fight": "URGENT", "intrusion": "URGENT", "weapon": "URGENT",
    "weapon_brandished": "URGENT", "fall": "URGENT", "trespass": "URGENT",
    "fire": "URGENT", "smoke": "URGENT", "shrinkage": "URGENT",
    "staff_zone":         "URGENT",     # default; per-rule override below
    "uniform_compliance": "ATTENTION", "shutter": "ATTENTION",
    "queue": "ATTENTION", "queue_length": "ATTENTION",
    "staff_present": "ATTENTION", "crowd": "ATTENTION",
    "abandoned_object": "ATTENTION", "loitering": "ATTENTION",
    "tailgating": "ATTENTION", "camera_offline": "ATTENTION",
}

# Four-tier severity ladder (spec Part 1 §1):
#   CRITICAL — immediate action required (theft, fight, weapon, fire,
#              fall, smoke, shrinkage, after-hours intrusion).
#   HIGH     — act within 5 minutes (suspicious behaviour, restricted
#              area, counter unstaffed, person after-hours).
#   MEDIUM   — review within 30 minutes (queue, loitering, uniform,
#              tailgating, crowd, sales-floor unattended).
#   LOW      — review end of day (routine heartbeats, on-time shop
#              open / close, sales-floor insight).
_SEVERITY_4: dict[str, str] = {
    "fight":              "CRITICAL",
    "weapon":             "CRITICAL",
    "weapon_brandished":  "CRITICAL",
    "fire":               "CRITICAL",
    "smoke":              "CRITICAL",
    "fall":               "CRITICAL",
    "shrinkage":          "CRITICAL",
    "intrusion":          "CRITICAL",

    "trespass":           "HIGH",
    "staff_present":      "HIGH",       # counter unstaffed
    "staff_zone":         "HIGH",       # default; per-rule override below
    "abandoned_object":   "HIGH",
    "camera_offline":     "HIGH",

    "queue":              "MEDIUM",
    "queue_length":       "MEDIUM",
    "crowd":              "MEDIUM",
    "loitering":          "MEDIUM",
    "tailgating":         "MEDIUM",
    "uniform_compliance": "MEDIUM",
    "shutter":            "MEDIUM",

    "live_activity":      "MEDIUM",
    "shop_open_close":    "LOW",
    "sales_floor_insight":"LOW",
    "store_intelligence": "LOW",
    "entry_exit":         "LOW",
    "dwell":              "LOW",
    "passersby":          "LOW",
    "occupancy":          "LOW",
    "person":             "LOW",        # default; per-context override below
}

_SEVERITY_4_COLOR: dict[str, str] = {
    "CRITICAL": "#dc2626",      # red-600
    "HIGH":     "#ea580c",      # orange-600
    "MEDIUM":   "#ca8a04",      # yellow-600
    "LOW":      "#2563eb",      # blue-600
}
_SEVERITY_4_EMOJI: dict[str, str] = {
    "CRITICAL": "🔴", "HIGH": "🟠", "MEDIUM": "🟡", "LOW": "🔵",
}

# Feed-ordering rank buckets — derived from the 4-tier ladder so EVERY
# critical/high detection type surfaces to the top of the alerts feed.
# (The old hand-maintained _CRITICAL_TYPES / _WARNING_TYPES lists silently
# dropped types like weapon_brandished / shrinkage / abandoned_object that
# were added to _SEVERITY_4 later.)
_RANK_CRITICAL = [dt for dt, s in _SEVERITY_4.items() if s == "CRITICAL"]
_RANK_HIGH     = [dt for dt, s in _SEVERITY_4.items() if s == "HIGH"]


def _severity_4_label(detection_type: str | None,
                      event: DetectionEvent | None = None,
                      zone: Zone | None = None, store=None) -> str:
    """4-tier severity ladder. Per-rule overrides for the detectors
    whose level depends on the event context."""
    dt = detection_type or ""
    extra = (event.extra or {}) if event is not None else {}
    rule = extra.get("rule", "")
    # Person: customer = LOW; after-hours / restricted = HIGH.
    if dt == "person":
        if event is not None:
            ctxt = _person_context(event, zone, store)
            return "LOW" if ctxt == "customer" else "HIGH"
        return "HIGH"
    # Uniform compliance: no-lanyard is gentle; wrong colour is louder.
    if dt == "uniform_compliance":
        return "LOW" if rule == "no_lanyard" else "MEDIUM"
    # Staff zone: customer/intruder behind counter is HIGH; missing
    # nametag is just LOW.
    if dt == "staff_zone":
        return "LOW" if rule == "missing_nametag" else "HIGH"
    # Sales-floor insight: low engagement / unattended floor = MEDIUM;
    # everything else (quiet/good/baseline) is the LOW heartbeat.
    if dt == "sales_floor_insight":
        return "MEDIUM" if rule in ("low_engagement", "unattended_floor") else "LOW"
    # Shop open/close: not-opened-by-cutoff is CRITICAL; before-hours
    # Live Activity Sentinel: severity rides on the rule.
    if dt == "live_activity":
        return {"after_hours_activity": "HIGH",
                "occupancy_surge":      "MEDIUM",
                "store_surge":          "MEDIUM",
                "dead_scene":           "MEDIUM",
                "activity_presence":    "LOW"}.get(rule, "MEDIUM")
    # / late-opening are HIGH/MEDIUM; routine open + close are LOW.
    if dt == "shop_open_close":
        if rule == "shop_not_opened":           return "CRITICAL"
        if rule == "shop_opened_before_hours":  return "HIGH"
        if rule == "shop_opened_late":          return "MEDIUM"
        return "LOW"
    return _SEVERITY_4.get(dt, "LOW")


def _severity_4_color(label: str) -> str:
    return _SEVERITY_4_COLOR.get(label, "#64748b")


def _severity_label(detection_type: str | None,
                    event: DetectionEvent | None = None,
                    zone: Zone | None = None, store=None) -> str:
    # Person detection is context-aware: customer = INFO, after-hours
    # or restricted-zone = URGENT. When called without context (e.g.
    # the summary count), a persisted person ALERT is always URGENT —
    # the worker only creates one when after-hours or restricted.
    if detection_type == "person":
        if event is not None:
            ctxt = _person_context(event, zone, store)
            return "INFO" if ctxt == "customer" else "URGENT"
        return "URGENT"
    # Uniform compliance: a missing name tag is just INFO (gentle
    # nudge); a person at the counter with no uniform at all is
    # ATTENTION (could be an unidentified person).
    if detection_type == "uniform_compliance" and event is not None:
        return "INFO" if (event.extra or {}).get("rule") == "no_lanyard" else "ATTENTION"
    # Staff-only zone: missing name tag is INFO, anything else
    # (unauthorised / customer in staff area) is URGENT.
    if detection_type == "staff_zone" and event is not None:
        return "INFO" if (event.extra or {}).get("rule") == "missing_nametag" else "URGENT"
    # Shop open / close: routine open + close are INFO; before-hours
    # is URGENT (security implication); late-opening is ATTENTION
    # (staffing issue, not an emergency).
    if detection_type == "live_activity":
        rule = (event.extra or {}).get("rule", "") if event is not None else ""
        if rule == "after_hours_activity":
            return "URGENT"
        return "INFO" if rule == "activity_presence" else "ATTENTION"
    if detection_type == "shop_open_close" and event is not None:
        rule = (event.extra or {}).get("rule", "")
        if rule in ("shop_opened_before_hours", "shop_not_opened"):
            return "URGENT"
        if rule == "shop_opened_late":
            return "ATTENTION"
        return "INFO"      # shop_opened, shop_closed
    # Sales-floor insight: heartbeat is INFO; low engagement and
    # unattended floor are ATTENTION (a manager-actionable nudge);
    # detection_offline is ATTENTION too — it's an ops / IT issue,
    # not a customer-flow signal.
    if detection_type == "sales_floor_insight" and event is not None:
        rule = (event.extra or {}).get("rule", "")
        if rule in ("low_engagement", "unattended_floor", "detection_offline"):
            return "ATTENTION"
        return "INFO"
    return _SEVERITY_LABEL.get(detection_type or "", "INFO")


# Plain-English card heading (no camera suffix) — the big title a
# non-technical manager reads first. Mirrors spec Part 2.
_PLAIN_TITLES: dict[str, str] = {
    "uniform_compliance": "Staff Not in Uniform",
    "intrusion":          "Someone in Store After Hours",
    "fight":              "Incident Detected — Check Immediately",
    "crowd":              "Too Many People in One Area",
    "trespass":           "Unauthorised Person in Staff Area",
    "shrinkage":          "Suspicious Activity Near Products",
    "fall":               "Person May Have Fallen — Check Now",
    "queue":              "Long Queue at Checkout",
    "queue_length":       "Long Queue at Checkout",
    "staff_present":      "Counter Left Unattended",
    "abandoned_object":   "Unattended Item Found",
    "camera_offline":     "Camera Not Working",
    "loitering":          "Person Lingering in One Area",
    "weapon":             "Weapon Detected — Call Security",
    "weapon_brandished":  "Weapon Detected — Call Security",
    "fire":               "Possible Fire — Check Now",
    "smoke":              "Possible Smoke — Check Now",
    "tailgating":         "Two People Entered Together",
}


def _plain_title(event: DetectionEvent, zone: Zone | None = None, store=None) -> str:
    dt = event.detection_type or "alert"
    extra = event.extra or {}
    if dt == "person":
        ctxt = _person_context(event, zone, store)
        if ctxt == "restricted":
            return "Unauthorised Person in Staff Area"
        if ctxt == "after_hours":
            return "Person Detected After Hours"
        return "Customer in Store"
    if dt == "shutter":
        rule = extra.get("rule", "")
        state = extra.get("shutter_state", "")
        if rule == "still_closed_at_opening" or state == "closed":
            return "Store Door Still Closed"
        if rule == "open_after_hours" or state == "open":
            return "Store Door Open After Hours"
        if rule == "partial_stuck":
            return "Store Door Stuck Part-Open"
        return "Store Door Issue"
    if dt == "shop_open_close":
        rule = extra.get("rule", "")
        eat = extra.get("eat_time", "")
        if rule == "shop_opened_before_hours":
            return f"⚠️ Store opened before trading hours ({eat})"
        if rule == "shop_opened_late":
            return f"⚠️ Store Opened Late ({eat})"
        if rule == "shop_opened":
            return f"✅ Store Opened ({eat})"
        if rule == "shop_opened_inferred":
            opened = extra.get("opened_at_eat") or eat
            return f"✅ Store Opened — inferred ({opened})"
        if rule == "shop_not_opened":
            return "🚨 Store Not Opened"
        if rule == "shop_closed":
            return f"✅ Store Closed ({eat})"
        if rule == "shop_daily_summary":
            return "📋 Daily Open/Close Summary"
        return "Shop open/close event"
    if dt == "store_intelligence":
        # Worker sets the full "Store Update — {store} — {time}" title.
        return extra.get("title") or (
            f"Store Update — {extra.get('store_name') or (store.name if store else 'Store')}")
    if dt == "sales_floor_insight":
        # All sales-floor insight rules now share one neutral title;
        # the severity colour + body carry the rule-specific tone.
        store_name = extra.get("store_name") or (store.name if store else "Store")
        return f"Status Update — {store_name}"
    if dt == "uniform_compliance":
        rule = extra.get("rule", "")
        if rule == "no_lanyard":
            return "Staff Missing Name Tag"
        # No uniform at all on a person at the counter — could be an
        # unidentified person, not just an out-of-uniform staffer.
        return "Unidentified Person at Service Counter"
    if dt == "staff_zone":
        rule = extra.get("rule", "")
        if rule == "customer_in_staff_zone":
            return "Customer in Staff-Only Area"
        if rule == "missing_nametag":
            return "Staff Member Missing Name Tag"
        return "Unidentified Person Behind Counter"
    if dt == "checkout_dwell":
        # All variants today are "long session" alerts. If the schema
        # grows more rules (e.g. abandoned-basket), branch on
        # extra.rule here.
        return "Checkout Taking Too Long"
    if dt == "intrusion":
        # time_context (Aug 2026) distinguishes a 07:30 presence from a
        # 21:30 one; older events without the field keep the after-
        # hours wording they always had.
        if extra.get("time_context") == "before_hours":
            return "Someone in Store Before Hours"
        return "Someone in Store After Hours"
    return _PLAIN_TITLES.get(dt, dt.replace("_", " ").title())


# "What to do" steps per alert type — max 3, plain action words.
# {placeholders} fill from settings so head office can wire real
# numbers without code changes.
_WHAT_TO_DO: dict[str, list[str]] = {
    "intrusion": ["Check the live camera now",
                  "Call building security: {security_phone}",
                  "Do not enter the store alone"],
    "trespass": ["Check the live camera now",
                 "Call building security: {security_phone}",
                 "Do not approach alone"],
    "uniform_compliance": ["Remind the staff member to wear their uniform and name tag",
                           "Check if it is a new staff member needing a uniform",
                           "Mark as resolved once sorted"],
    "shutter": ["Call the store: {store_phone}",
                "Check if staff are on their way",
                "Mark resolved when the store opens"],
    "queue": ["Open a second till if available",
              "Call a free staff member to help",
              "Let waiting customers know"],
    "queue_length": ["Open a second till if available",
                     "Call a free staff member to help",
                     "Let waiting customers know"],
    "staff_present": ["Ask nearby staff to cover the counter",
                      "Check if the staff member is on a break",
                      "Make sure the counter is always covered"],
    "camera_offline": ["Check the camera power cable is connected",
                       "Restart the camera from the NVR",
                       "Call IT support if still offline: {it_phone}"],
    "fight": ["Check the live camera now",
              "Send staff to the scene",
              "Call building security: {security_phone}"],
    "fall": ["Check the live camera now",
             "Send help to the person",
             "Call an ambulance if they are hurt"],
    "crowd": ["Check the area on the live camera",
              "Send a staff member to manage the crowd",
              "Watch for safety and blocked exits"],
    "shrinkage": ["Review the footage",
                  "Send a staff member to the area discreetly",
                  "Follow your loss-prevention steps"],
    "abandoned_object": ["Check the live camera",
                         "Send a staff member to inspect the item",
                         "Call security if it looks suspicious"],
    "weapon": ["Call building security immediately: {security_phone}",
               "Keep staff and customers away from the area",
               "Do not approach"],
    "weapon_brandished": ["Call building security immediately: {security_phone}",
                          "Keep staff and customers away from the area",
                          "Do not approach"],
}


def _what_to_do(event: DetectionEvent, store, zone: Zone | None = None) -> list[str]:
    dt = event.detection_type or ""
    if dt == "checkout_dwell":
        # The body lists the three plausible causes inline; an
        # additional "what to do" list would be redundant noise.
        return []
    if dt == "person":
        ctxt = _person_context(event, zone, store)
        if ctxt == "customer":
            return []   # no action — it's just a customer
        if ctxt == "restricted":
            steps = ["Check who is in the staff area",
                     "Ask the person to leave if unauthorised",
                     "Mark resolved when the area is clear"]
        else:  # after_hours
            steps = ["Check the live camera now",
                     "Call building security: {security_phone}",
                     "Do not enter the store alone"]
    elif dt == "uniform_compliance":
        if (event.extra or {}).get("rule") == "no_lanyard":
            steps = ["Remind the staff member to wear their name tag",
                     "Check if it is a new staff member needing one",
                     "Mark resolved once sorted"]
        else:
            steps = ["Check the live camera now",
                     "Confirm whether they are a staff member",
                     "Ask them to leave the counter if unauthorised"]
    elif dt == "staff_zone":
        rule = (event.extra or {}).get("rule", "")
        if rule == "customer_in_staff_zone":
            steps = ["Check the live camera now",
                     "Politely guide the customer back to the customer area"]
        elif rule == "missing_nametag":
            steps = ["Remind staff to wear their name tag",
                     "Check if it is a new staff member"]
        else:
            steps = ["Check the live camera now",
                     "Ask the person to identify themselves",
                     "Guide customers back to the customer area if needed"]
    elif dt == "sales_floor_insight":
        rule = (event.extra or {}).get("rule", "")
        if rule == "unattended_floor":
            steps = ["Send a staff member to the sales floor",
                     "Greet customers and offer assistance",
                     "Mark resolved once staff are in position"]
        elif rule == "low_engagement":
            steps = ["Check if popular displays are well stocked",
                     "Move popular items to more visible locations",
                     "Mark resolved when adjusted"]
        elif rule == "good_engagement":
            steps = ["Keep popular zones well stocked",
                     "No action needed — mark resolved"]
        elif rule == "detection_offline":
            steps = ["Check the live camera view if anything looks off",
                     "Mark resolved once detection data resumes"]
        elif rule == "quiet_period":
            steps = ["No action needed — this may be normal for the time of day",
                     "Mark resolved"]
        else:
            steps = ["Check if popular zones are well stocked",
                     "Ensure staff are available in busy areas",
                     "No action needed if all looks good"]
    elif dt == "shop_open_close":
        rule = (event.extra or {}).get("rule", "")
        if rule == "shop_opened_before_hours":
            steps = ["Check the live camera now",
                     "Confirm whether opening early is authorised",
                     "Call the store manager if unexpected: {store_phone}"]
        elif rule == "shop_opened_late":
            steps = ["Note the late opening time",
                     "Check why opening was delayed",
                     "Mark resolved once trading has begun"]
        elif rule == "shop_opened":
            steps = ["Routine opening — no action needed",
                     "Mark resolved"]
        elif rule == "shop_opened_inferred":
            steps = ["Routine opening — no action needed",
                     "Check the entry/exit line setup if this keeps inferring instead of recording a crossing",
                     "Mark resolved"]
        elif rule == "shop_not_opened":
            steps = ["Call the store manager: {store_phone}",
                     "Check the live camera at the entrance",
                     "Mark resolved once the store has opened"]
        elif rule == "shop_closed":
            steps = ["Routine closing — no action needed",
                     "Mark resolved"]
        elif rule == "shop_daily_summary":
            steps = ["Daily summary — no action needed",
                     "Review the open / close times if anything looks off",
                     "Mark resolved"]
        else:
            steps = ["Check the live camera"]
    else:
        steps = _WHAT_TO_DO.get(dt)
    if not steps:
        return ["Check the live camera",
                "Decide whether action is needed",
                "Mark resolved when handled"]
    store_phone = (getattr(store, "manager_phone", None)
                   or getattr(settings, "store_default_phone", "") or "the store")
    security_phone = getattr(settings, "security_phone", "") or "building security"
    it_phone = getattr(settings, "it_support_phone", "") or "IT support"
    return [s.format(store_phone=store_phone, security_phone=security_phone,
                     it_phone=it_phone) for s in steps]


# Per-type emoji prefix for the title — the spec called these out
# explicitly. Operators scan the feed quickly; the icon does the
# heavy lifting for category recognition.
_TITLE_ICONS: dict[str, str] = {
    "fight":             "🚨",
    "intrusion":         "🚨",
    "trespass":          "🚨",
    "weapon":            "🚨",
    "weapon_brandished": "🚨",
    "fall":              "🚨",
    "fire":              "🔥",
    "smoke":             "💨",
    "shrinkage":         "⚠️",
    "queue":             "⏱️",
    "queue_length":      "⏱️",
    "crowd":             "⚠️",
    "loitering":         "⚠️",
    "shutter":           "🔒",
    "live_activity":     "👥",
    "shop_open_close":   "🏬",
    "sales_floor_insight": "📊",
    "store_intelligence":  "📊",
    "abandoned_object":  "🧳",
    "tailgating":        "⚠️",
    "staff_present":     "👤",
    "occupancy":         "📊",
}


def _icon(detection_type: str | None) -> str:
    return _TITLE_ICONS.get(detection_type or "", "•")


def _extract(extra: dict | None, *keys, default=None):
    """Best-effort getter across an alert's `event.extra` blob — the
    detector usually drops the contextual numbers in there
    (`duration_min`, `count`, `wait_seconds`, etc)."""
    if not isinstance(extra, dict):
        return default
    for k in keys:
        if k in extra and extra[k] is not None:
            return extra[k]
    return default


# Zone tags that make a person detection "restricted" (staff-only).
_PERSON_RESTRICTED_TAGS = {"restricted", "intrusion", "trespass", "high_value", "stockroom"}


def _zone_restricted(zone: Zone | None) -> bool:
    if zone is None:
        return False
    return bool(_PERSON_RESTRICTED_TAGS & set(zone.detection_types_json or []))


def _is_after_hours(event: DetectionEvent, store) -> bool:
    """True if the event happened outside the store's business hours
    AND outside the configurable pre-opening / post-closing grace
    windows.

    Mirrors the worker-side gate so the alert label ("Person Detected
    After Hours") only attaches to events the worker also classified
    as after-hours. Time math uses Africa/Nairobi via
    is_after_hours_with_grace().
    """
    if store is None:
        return False   # unknown store → don't cry wolf
    try:
        from app.utils.business_hours import is_after_hours_with_grace
        from app.config import settings
        ts = event.timestamp
        if ts is not None and ts.tzinfo is None:
            from datetime import timezone as _tz
            ts = ts.replace(tzinfo=_tz.utc)
        return is_after_hours_with_grace(
            store, ts,
            grace_before_open_min=settings.person_afterhours_grace_before_min,
            grace_after_close_min=settings.person_afterhours_grace_after_min,
        )
    except Exception:
        return False


def _person_context(event: DetectionEvent, zone: Zone | None, store) -> str:
    """Classify a person detection: 'restricted' | 'after_hours' |
    'customer'. Drives the context-aware title/body/severity."""
    if _zone_restricted(zone):
        return "restricted"
    if _is_after_hours(event, store):
        return "after_hours"
    return "customer"


def _title(event: DetectionEvent, camera: Camera | None,
           zone: Zone | None = None, store=None) -> str:
    """Human-readable title for an alert. Embeds the camera name so
    operators triaging a long feed know WHICH camera fired."""
    cam = camera.name if camera else "unknown camera"
    dt = event.detection_type or "alert"
    extra = event.extra or {}
    icon = _icon(dt)

    if dt == "person":
        ctxt = _person_context(event, zone, store)
        if ctxt == "restricted":
            return f"🚨 Unauthorised Person in Staff Area — {cam}"
        if ctxt == "after_hours":
            return f"🚨 Person Detected After Hours — {cam}"
        return f"👤 Customer in Store — {cam}"

    if dt == "staff_present":
        mins = _extract(extra, "unstaffed_minutes", "duration_min", default=None)
        if mins is not None:
            return f"{icon} Counter Unstaffed for {int(round(float(mins)))} min — {cam}"
        return f"{icon} Counter Unstaffed — {cam}"
    if dt in ("queue", "queue_length"):
        n = _extract(extra, "queue_length", "count", default=None)
        if n is not None:
            return f"{icon} Long Queue — {int(n)} people waiting — {cam}"
        return f"{icon} Long Queue — {cam}"
    if dt == "intrusion":
        if extra.get("time_context") == "before_hours":
            return f"{icon} Before-hours Intrusion Detected — {cam}"
        return f"{icon} After-hours Intrusion Detected — {cam}"
    if dt == "crowd":
        n = _extract(extra, "count", "people", default=None)
        if n is not None:
            return f"{icon} Crowd Alert — {int(n)} people in zone — {cam}"
        return f"{icon} Crowd Alert — {cam}"
    if dt == "shutter":
        state = _extract(extra, "state", "shutter_state", default="")
        return f"{icon} Shutter {state or 'state change'} — {cam}"
    if dt == "live_activity":
        rule = extra.get("rule", "")
        if rule == "activity_presence":
            n = extra.get("people_count")
            count = (f" ({int(n)} people)"
                     if isinstance(n, (int, float)) else "")
            return f"👁️ Live Activity Presence{count} — {cam}"
        label = {
            "occupancy_surge":      "Occupancy Surge",
            "store_surge":          "Store-Wide Occupancy Surge",
            "after_hours_activity": "After-Hours Activity Detected",
            "dead_scene":           "Camera Activity Stalled",
        }.get(rule, "Live Activity")
        n = extra.get("people_count")
        count = (f" ({int(n)} people)"
                 if isinstance(n, (int, float)) and rule != "dead_scene" else "")
        return f"👥 {label}{count} — {cam}"
    if dt == "shop_open_close":
        rule = extra.get("rule", "")
        eat = extra.get("eat_time", "")
        store_name = extra.get("store_name") or (store.name if store else "Store")
        if rule == "shop_opened_before_hours":
            return f"⚠️ {store_name} opened before trading hours ({eat})"
        if rule == "shop_opened_late":
            return f"⚠️ Store Opened Late — {store_name} ({eat})"
        if rule == "shop_opened":
            return f"✅ Store Opened — {store_name} ({eat})"
        if rule == "shop_opened_inferred":
            opened = extra.get("opened_at_eat") or eat
            return f"✅ Store Opened — {store_name} (inferred {opened})"
        if rule == "shop_not_opened":
            return f"🚨 Store Not Opened — {store_name}"
        if rule == "shop_closed":
            return f"✅ Store Closed — {store_name} ({eat})"
        if rule == "shop_daily_summary":
            return f"📋 Daily Summary — {store_name}"
        return f"{icon} Shop open/close — {cam}"
    if dt in ("sales_floor_insight", "store_intelligence"):
        # Store-scoped, not camera-scoped — the camera-suffix would be noise
        # for managers. Pull straight from the plain-title path.
        return _plain_title(event, zone, store)
    if dt == "checkout_dwell":
        # Also store-scoped — the manager doesn't care which till
        # camera; they care which store has a stuck checkout.
        store_name = (store.name if store else None) \
                      or (extra.get("store_name")) or "Unknown store"
        return f"{icon} Checkout Taking Too Long — {store_name}"
    if dt == "trespass":
        return f"{icon} Unauthorised person in restricted zone — {cam}"
    if dt == "fight":
        return f"{icon} Incident / fight detected — {cam}"
    if dt == "occupancy":
        pct = _extract(extra, "capacity_pct", default=None)
        if pct is not None:
            return f"{icon} Store at {int(round(float(pct)))}% capacity — {cam}"
        return f"{icon} Occupancy alert — {cam}"
    if dt == "shrinkage":
        return f"{icon} Potential loss-prevention alert — {cam}"
    if dt == "fall":
        return f"{icon} Person fall detected — {cam}"
    if dt == "tailgating":
        return f"{icon} Tailgate detected — {cam}"
    if dt == "abandoned_object":
        return f"{icon} Abandoned item — {cam}"
    if dt == "loitering":
        mins = _extract(extra, "duration_min", default=None)
        if mins is not None:
            return f"{icon} Loitering for {int(round(float(mins)))} min — {cam}"
        return f"{icon} Loitering — {cam}"

    # Fallback: Titlecase the snake_case detection type.
    label = dt.replace("_", " ").title()
    return f"{icon} {label} — {cam}"


def _body(event: DetectionEvent, zone: Zone | None, store=None) -> str:
    """One-sentence description with the relevant operational
    implication. These are the bits store managers actually need to
    decide whether to act."""
    dt = event.detection_type or ""
    extra = event.extra or {}
    zone_name = zone.name if zone else None

    if dt == "person":
        ctxt = _person_context(event, zone, store)
        when = ""
        try:
            if event.timestamp is not None:
                when = " at " + event.timestamp.strftime("%I:%M %p").lstrip("0")
        except Exception:
            pass
        store_name = (getattr(store, "name", None) or "the store")
        if ctxt == "restricted":
            return (f"Someone was detected in a staff-only area{when}. "
                    f"This area is restricted to staff members only.")
        if ctxt == "after_hours":
            return (f"Someone was seen inside {store_name}{when}, "
                    f"outside opening hours. Please check immediately.")
        return (f"A customer was detected in {store_name}{when}. "
                f"This is normal during business hours.")

    if dt == "uniform_compliance":
        store_name = (getattr(store, "name", None) or "the store")
        if extra.get("rule") == "no_lanyard":
            return (f"A staff member at the {store_name} counter has been "
                    f"without a visible lanyard or name tag. Please remind "
                    f"them to wear it.")
        return (f"A person at the {store_name} service counter is not in "
                f"uniform. Check whether they are a new staff member or "
                f"someone who shouldn't be behind the counter.")

    if dt == "staff_zone":
        store_name = (getattr(store, "name", None) or "the store")
        mins = _extract(extra, "duration_minutes", default=None)
        secs = _extract(extra, "duration_seconds", default=None)
        when = (f"for {int(mins)} minute{'s' if int(mins) != 1 else ''}"
                if mins and int(mins) >= 1
                else f"for {int(secs)} seconds" if secs else "")
        rule = extra.get("rule", "")
        if rule == "customer_in_staff_zone":
            return (f"A customer has entered the staff area at {store_name}. "
                    f"Please guide them back to the customer area.")
        if rule == "missing_nametag":
            return (f"A staff member at {store_name} has been working without "
                    f"a visible name tag {when}.").strip()
        return (f"Someone without a staff uniform has been behind the "
                f"service counter at {store_name} {when}. Please check "
                f"immediately.").strip()

    if dt == "staff_present":
        mins = _extract(extra, "unstaffed_minutes", "duration_min", default=None)
        zone = extra.get("zone_name") or "service"
        last = extra.get("last_activity_eat")
        if mins is not None:
            base = (f"No person detected at the {zone} counter for the past "
                    f"{int(round(float(mins)))} minutes.")
        else:
            base = f"No person detected at the {zone} counter."
        if last:
            base += f" Last activity: {last}."
        else:
            base += " Customer service may be affected."
        return base
    if dt in ("queue", "queue_length"):
        n = _extract(extra, "queue_length", "count", default=None)
        wait = _extract(extra, "wait_seconds", "queue_wait_seconds", default=None)
        sentences = []
        if n is not None:
            sentences.append(f"Queue has grown to {int(n)} people.")
        if wait is not None:
            sentences.append(f"Average wait time is estimated at {int(round(float(wait) / 60))} minutes.")
        sentences.append("Consider opening an additional checkout point.")
        return " ".join(sentences)
    if dt == "intrusion":
        return ("Motion detected inside the store outside business hours. "
                "Verify whether this is authorised staff or an actual intrusion.")
    if dt == "crowd":
        n = _extract(extra, "count", "people", default=None)
        if n is not None:
            return (f"{int(n)} people detected in {zone_name or 'a single zone'} simultaneously. "
                    f"Monitor for safety and check for bottlenecks.")
        return ("Large gathering detected. Monitor for safety.")
    if dt == "shutter":
        return ("Shutter open/close state has changed. Verify against expected store hours.")
    if dt == "live_activity":
        rule = extra.get("rule", "")
        if rule == "dead_scene":
            return (extra.get("message")
                    or "Camera is streaming but produced no detections.")
        cam_name = extra.get("camera_name") or "camera"
        people = int(extra.get("people_count") or 0)
        suffix = {"occupancy_surge":      " (surge)",
                  "store_surge":          " (store-wide surge)",
                  "after_hours_activity": " after hours"}.get(rule, "")
        parts = [f"{people} people detected at {cam_name}{suffix}"]
        staff = extra.get("staff_count")
        customers = extra.get("customer_count")
        if isinstance(staff, int) and isinstance(customers, int) \
                and (staff > 0 or customers > 0):
            breakdown = []
            if staff > 0:
                breakdown.append(f"{staff} staff")
            if customers > 0:
                breakdown.append(
                    f"{customers} customer{'s' if customers != 1 else ''}")
            parts.append(f"({', '.join(breakdown)})")
        return " | ".join(parts)
    if dt == "shop_open_close":
        msg = extra.get("message") or ""
        if msg:
            return msg
        rule = extra.get("rule", "")
        eat = extra.get("eat_time", "")
        if rule == "shop_opened_before_hours":
            return f"Shop opened before trading hours at {eat}."
        if rule == "shop_opened_late":
            return f"Shop opened late at {eat}."
        if rule == "shop_opened":
            return f"Shop open at {eat}."
        if rule == "shop_closed":
            return f"Shop closed at {eat}."
        return "Shop open/close event recorded."
    if dt == "store_intelligence":
        # Worker pre-formats the full multi-section BI body.
        return extra.get("message") or "Store intelligence update."
    if dt == "sales_floor_insight":
        # Worker pre-formats the body; pass it through. Falls back to
        # a short summary if message is missing (older row).
        msg = extra.get("message") or ""
        if msg:
            return msg
        n = extra.get("total_customers")
        avg = extra.get("avg_browse_seconds")
        if n is not None and avg is not None:
            return (f"{n} customers browsed the sales floor in the last "
                    f"15 minutes. Average browse time {avg:.0f}s.")
        return "Sales floor activity summary."
    if dt == "trespass":
        return (f"An unauthorised person was detected in {zone_name or 'a restricted area'}. "
                f"Investigate and respond per protocol.")
    if dt == "fight":
        return ("Aggressive or fighting behaviour detected. Dispatch staff to scene immediately.")
    if dt == "fall":
        return ("A person appears to have fallen. Send help and check for injury.")
    if dt == "occupancy":
        pct = _extract(extra, "capacity_pct", default=None)
        if pct is not None:
            return (f"Store occupancy is at {int(round(float(pct)))}% of stated capacity. "
                    f"Watch for crowding at peak.")
        return ("Occupancy threshold reached.")
    if dt == "shrinkage":
        return ("Behaviour consistent with potential shoplifting detected. "
                "Review footage and dispatch staff if appropriate.")
    if dt == "tailgating":
        return ("Multiple people passed through a single-entry zone together. "
                "Verify each is authorised.")
    if dt == "abandoned_object":
        return ("An object has been left unattended in the store. Investigate.")
    if dt == "loitering":
        return ("A person has been lingering in one area beyond the typical browsing window.")
    if dt == "checkout_dwell":
        # `dwell_seconds` is set by the alerting task. Format as
        # "X minutes Y seconds" with the noun matched to the value
        # so "1 minute 1 second" reads correctly when relevant.
        extra = event.extra or {}
        secs = float(extra.get("dwell_seconds") or 0)
        m, s = divmod(int(round(secs)), 60)
        parts = []
        if m:
            parts.append(f"{m} minute{'s' if m != 1 else ''}")
        parts.append(f"{s} second{'s' if s != 1 else ''}")
        duration_str = " ".join(parts) if parts else "an extended period"

        store_name = (store.name if store else None) \
                      or extra.get("store_name") or "the store"
        return (
            f"Someone has been at the till at {store_name} for "
            f"{duration_str}. Open the live camera to check and "
            f"assist if needed."
        )

    return "Review the footage and decide whether to confirm or dismiss."


def _snapshot_url(alert_id: int, event: DetectionEvent) -> str:
    """Browser-fetchable URL for an alert's snapshot. Always returns
    a URL — the endpoint itself falls back to the camera's latest
    cached frame when the event has no archived thumbnail. Frontend
    can <img src> this directly; on 404 it renders the camera-icon
    placeholder."""
    return f"/api/alerts/{alert_id}/snapshot"


def _time_range(event: DetectionEvent, store) -> str | None:
    """Human-readable when-it-happened line for the alert card.

    Duration-style events (dwell, staff_present, loitering,
    abandoned_object) show a "From 9:30 PM to 9:55 PM" range, computed
    from the event's known duration field. Point-in-time events
    (intrusion, fight, queue, crowd) just show "At 9:30 PM".

    Timestamps are formatted in the STORE's timezone so a Junction
    manager sees 9:30 PM Nairobi, a future Kampala manager sees
    9:30 PM EAT, etc.
    """
    if not event or not event.timestamp:
        return None
    try:
        from zoneinfo import ZoneInfo
        tz_name = (getattr(store, "timezone", None) or "Africa/Nairobi") if store else "Africa/Nairobi"
        try:
            tz = ZoneInfo(tz_name)
        except Exception:
            tz = ZoneInfo("Africa/Nairobi")
    except ImportError:
        tz = None     # py<3.9 fallback — extremely unlikely on our image
    ts = event.timestamp
    if tz:
        ts_local = ts.astimezone(tz)
    else:
        ts_local = ts
    fmt = "%-I:%M %p"   # "9:30 PM" — POSIX strftime; Windows would need %#I
    try:
        end_str = ts_local.strftime(fmt)
    except Exception:
        # Windows compatibility (the Docker host is usually Linux but
        # guard anyway).
        end_str = ts_local.strftime("%I:%M %p").lstrip("0")

    # Extract the per-event duration when the detector provided it.
    extra = event.extra or {}
    duration_seconds: float | None = None
    dt = event.detection_type or ""
    if dt == "staff_present":
        m = (extra.get("unstaffed_minutes") or extra.get("duration_min")
             or extra.get("duration_minutes"))
        if m is not None:
            try:
                duration_seconds = float(m) * 60.0
            except Exception:
                duration_seconds = None
    elif dt == "dwell":
        s = (extra.get("dwell_seconds") or extra.get("duration_seconds")
             or extra.get("duration_sec"))
        if s is not None:
            try:
                duration_seconds = float(s)
            except Exception:
                duration_seconds = None
    elif dt == "loitering":
        m = extra.get("duration_min") or extra.get("duration_minutes")
        if m is not None:
            try:
                duration_seconds = float(m) * 60.0
            except Exception:
                duration_seconds = None
    elif dt == "abandoned_object":
        s = extra.get("duration_seconds") or extra.get("dwell_seconds")
        if s is not None:
            try:
                duration_seconds = float(s)
            except Exception:
                duration_seconds = None

    if duration_seconds and duration_seconds >= 60:
        from datetime import timedelta
        start_ts = ts_local - timedelta(seconds=duration_seconds)
        try:
            start_str = start_ts.strftime(fmt)
        except Exception:
            start_str = start_ts.strftime("%I:%M %p").lstrip("0")
        mins = int(round(duration_seconds / 60))
        return f"🕒 Between {start_str} and {end_str} ({mins} min)"
    # Point-in-time event.
    return f"🕒 Detected at {end_str}"


def _to_alert_out(alert: Alert, event: DetectionEvent,
                  camera: Camera | None, zone: Zone | None,
                  store=None) -> AlertOut:
    item = AlertOut.model_validate(alert)
    item.camera_id      = event.camera_id
    item.camera_name    = camera.name if camera else None
    item.detection_type = event.detection_type
    item.confidence     = event.confidence
    item.bbox_norm      = event.bbox_json
    item.zone_id        = event.zone_id
    item.zone_name      = zone.name if zone else None
    item.thumbnail_path = event.thumbnail_path
    item.severity       = _severity(event.detection_type)
    item.severity_label = _severity_label(event.detection_type, event, zone, store)
    # Four-tier ladder for the redesigned alerts page. severity_label
    # (3-tier URGENT/ATTENTION/INFO) stays for backwards-compat with
    # any cached client bundles.
    s4 = _severity_4_label(event.detection_type, event, zone, store)
    item.severity_4       = s4
    item.severity_4_color = _severity_4_color(s4)
    item.severity_4_emoji = _SEVERITY_4_EMOJI.get(s4, "•")
    item.title          = _title(event, camera, zone, store)
    item.plain_title    = _plain_title(event, zone, store)
    item.body           = _body(event, zone, store)
    item.what_to_do     = _what_to_do(event, store, zone)
    item.time_range     = _time_range(event, store)
    item.snapshot_url   = _snapshot_url(alert.id, event)
    # Checkout-dwell timeline (NULL for every other alert type).
    # Pass the raw path list AND a count so the frontend can render
    # the filmstrip without a separate length query.
    sp = list(alert.snapshot_paths or [])
    item.snapshot_paths  = sp or None
    item.snapshot_count  = len(sp) or None
    # Recorded clip (if the recorder extracted one for this alert). We only
    # expose the URL — the path stays server-side.
    item.clip_url = (f"/api/alerts/{alert.id}/clip"
                     if (event.extra or {}).get("alert_clip_path") else None)
    # VLM scene description lives in the event's extra JSON (written
    # async by the vlm.analyse_alert_scene task). Surface just that
    # one field — never the whole extra blob.
    item.vlm_scene = (event.extra or {}).get("vlm_scene") if event else None
    # Store Intelligence structured payload for the special metric-tile card.
    if event and event.detection_type == "store_intelligence":
        ex = event.extra or {}
        item.store_intel = {
            "store_name":      ex.get("store_name"),
            "city":            ex.get("city"),
            "time_eat":        ex.get("time_eat"),
            "time_period":     ex.get("time_period"),
            "people_count":    ex.get("people_count"),
            "staff_count":     ex.get("staff_count"),
            "counter_status":  ex.get("counter_status"),
            "busiest_zone":    ex.get("busiest_zone"),
            "entry_count_45m": ex.get("entry_count_45m"),
            "alert_count_45m": ex.get("alert_count_45m"),
            "hours_open":      ex.get("hours_open"),
            "ai_summary":      ex.get("ai_summary"),
            "recommendation":  ex.get("recommendation"),
        }
    return item


# ---- Endpoints -----------------------------------------------------

@router.get("/summary")
@cached_store_endpoint("alerts-summary", ttl=15)
def alerts_summary(db: Session = Depends(get_db),
                   _u: User = Depends(get_current_user),
                   store_id: Optional[int] = Query(None)):
    """Counts + lifecycle metrics for the alerts page's executive
    summary bar and the sidebar badge.

    Returns BOTH the legacy 3-tier counts (urgent / attention) and
    the new 4-tier counts (critical / high / medium / low) so older
    clients keep working while the redesigned page reads the new
    fields directly. Adds avg-time-to-resolve and a vs-yesterday
    trend so the bar can render "Avg Response: 1m 42s · 📈 Alerts
    up 12% vs yesterday" without a follow-up request."""
    from types import SimpleNamespace
    from app.models import Store as _Store
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    yest_start = today - timedelta(days=1)

    # Yesterday is only ever used as a COUNT for the trend line — one
    # aggregate query instead of materialising every ORM row.
    yq = (db.query(func.count(Alert.id))
            .join(DetectionEvent, Alert.event_id == DetectionEvent.id)
            .outerjoin(Camera, DetectionEvent.camera_id == Camera.id)
            .filter(DetectionEvent.timestamp >= yest_start,
                    DetectionEvent.timestamp < today))
    if store_id is not None:
        yq = yq.filter(Camera.store_id == store_id)
    yest_count = int(yq.scalar() or 0)

    # Today: the severity classifiers are rule-based Python over
    # event.extra + zone/store context, so classification stays in
    # Python — but on a COLUMN PROJECTION (7 columns + the tiny Store
    # entity) instead of full Alert+DetectionEvent ORM objects.
    tq = (db.query(Alert.status, Alert.created_at, Alert.resolved_at,
                   Alert.acknowledged_at,
                   DetectionEvent.detection_type, DetectionEvent.extra,
                   DetectionEvent.zone_id, DetectionEvent.timestamp,
                   _Store)
            .join(DetectionEvent, Alert.event_id == DetectionEvent.id)
            .outerjoin(Camera, DetectionEvent.camera_id == Camera.id)
            .outerjoin(_Store, Camera.store_id == _Store.id)
            .filter(DetectionEvent.timestamp >= today,
                    DetectionEvent.timestamp < today + timedelta(days=1)))
    if store_id is not None:
        tq = tq.filter(Camera.store_id == store_id)
    rows = tq.all()

    zone_ids = {r.zone_id for r in rows if r.zone_id is not None}
    zones_by_id = ({z.id: z for z in db.query(Zone).filter(Zone.id.in_(zone_ids)).all()}
                   if zone_ids else {})

    urgent = attention = resolved = dismissed = unread_urgent = 0
    critical = high = medium = low = 0
    resolve_durations: list[float] = []
    for r in rows:
        zone = zones_by_id.get(r.zone_id) if r.zone_id else None
        # Duck-typed event shim carrying the only fields the severity
        # classifiers (and their _person_context/_is_after_hours helpers)
        # read: extra, timestamp, detection_type.
        ev = SimpleNamespace(extra=r.extra, timestamp=r.timestamp,
                             detection_type=r.detection_type)
        store = r[8]
        label = _severity_label(r.detection_type, ev, zone, store)
        s4    = _severity_4_label(r.detection_type, ev, zone, store)
        if r.status in ("resolved", "confirmed"):
            resolved += 1
            # Time-to-resolve = resolved_at − created_at when both
            # present. Falls back to acknowledged_at when an older
            # row never recorded the resolve timestamp explicitly.
            r_at = r.resolved_at or r.acknowledged_at
            if r_at and r.created_at:
                if r_at.tzinfo is None: r_at = r_at.replace(tzinfo=timezone.utc)
                c_at = r.created_at
                if c_at.tzinfo is None: c_at = c_at.replace(tzinfo=timezone.utc)
                resolve_durations.append(max(0.0, (r_at - c_at).total_seconds()))
            continue
        if r.status == "dismissed":
            dismissed += 1
            continue
        if label == "URGENT":
            urgent += 1
            if r.status == "new":
                unread_urgent += 1
        elif label == "ATTENTION":
            attention += 1
        if s4 == "CRITICAL": critical += 1
        elif s4 == "HIGH":   high += 1
        elif s4 == "MEDIUM": medium += 1
        else:                low += 1

    avg_response_seconds = (sum(resolve_durations) / len(resolve_durations)
                            if resolve_durations else None)
    today_count = len(rows)
    trend_vs_yesterday_pct = None
    if yest_count > 0:
        trend_vs_yesterday_pct = round(
            (today_count - yest_count) / yest_count * 100.0, 1)
    elif today_count > 0:
        trend_vs_yesterday_pct = 100.0

    # Friendly date label in the store's timezone (or EAT default).
    from zoneinfo import ZoneInfo
    try:
        tz = ZoneInfo("Africa/Nairobi")
    except Exception:
        tz = timezone.utc
    date_label = datetime.now(tz).strftime("%A %-d %B %Y") \
        if hasattr(datetime, "strftime") else None
    # Some platforms (Windows) reject %-d — fall back to the padded form.
    if date_label is None or "-" in date_label:
        date_label = datetime.now(tz).strftime("%A %d %B %Y")

    return {
        # Legacy 3-tier — kept for backwards compat with cached bundles.
        "urgent": urgent, "attention": attention,
        "resolved_today": resolved,
        "dismissed_today": dismissed,
        "unread_urgent": unread_urgent,
        # New 4-tier ladder.
        "critical_today": critical,
        "high_today":     high,
        "medium_today":   medium,
        "low_today":      low,
        # Lifecycle stats.
        "avg_response_seconds":  avg_response_seconds,
        "today_count":           today_count,
        "yesterday_count":       yest_count,
        "trend_vs_yesterday_pct": trend_vs_yesterday_pct,
        "date_label":            date_label,
    }


# ---- Acknowledge endpoint ----------------------------------------

@router.post("/{alert_id}/acknowledge", response_model=AlertActionOut)
def acknowledge_alert(alert_id: int, db: Session = Depends(get_db),
                      _u: User = Depends(get_current_user)):
    """Mark the alert as acknowledged — drives the Generated →
    Acknowledged → Resolved progress bar. Idempotent: a second
    acknowledge on the same alert is a no-op."""
    alert = db.get(Alert, alert_id)
    if not alert:
        raise HTTPException(404, "alert not found")
    if alert.acknowledged_at is None:
        alert.acknowledged_at = datetime.now(timezone.utc)
        db.commit()
    return AlertActionOut(id=alert.id, status=alert.status)


@router.get("/export.xlsx")
def export_alerts_xlsx(
    db: Session = Depends(get_db),
    _u: User = Depends(get_current_user),
    store_id: Optional[int] = Query(None),
    detection_type: Optional[str] = Query(None),
    since: Optional[datetime] = Query(None),
    until: Optional[datetime] = Query(None),
):
    """Alert history as an .xlsx for manager review. Honours the same
    date / store / type filters as the list. Capped at 5000 rows so a
    careless 90-day pull doesn't OOM the API."""
    import io
    from fastapi.responses import StreamingResponse
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(503, "Excel export not available — rebuild the api image")

    from app.models import Store as _Store
    q = (db.query(Alert, DetectionEvent, Camera, _Store)
           .join(DetectionEvent, Alert.event_id == DetectionEvent.id)
           .outerjoin(Camera, DetectionEvent.camera_id == Camera.id)
           .outerjoin(_Store, Camera.store_id == _Store.id))
    if store_id is not None:
        q = q.filter(Camera.store_id == store_id)
    if detection_type:
        q = q.filter(DetectionEvent.detection_type == detection_type)
    if since:
        q = q.filter(DetectionEvent.timestamp >= since)
    if until:
        q = q.filter(DetectionEvent.timestamp <= until)
    rows = q.order_by(desc(Alert.created_at)).limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Alerts"
    headers = ["Date & Time", "Store", "Camera", "Alert", "Priority",
               "Status", "Notes"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1E293B")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
    for alert, ev, cam, store in rows:
        ts = ev.timestamp.strftime("%Y-%m-%d %H:%M") if ev.timestamp else ""
        ws.append([
            ts,
            store.name if store else "",
            cam.name if cam else "",
            _plain_title(ev),
            _severity_label(ev.detection_type),
            alert.status,
            (alert.notes or "").replace("\n", " | "),
        ])
    widths = [18, 18, 18, 32, 12, 12, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"vivoguard_alerts_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("", response_model=list[AlertOut])
def list_alerts(
    db: Session = Depends(get_db),
    _u: User = Depends(get_current_user),
    camera_id: Optional[int]   = Query(None),
    store_id: Optional[int]    = Query(None),
    detection_type: Optional[str] = Query(None),
    zone_id: Optional[int]     = Query(None),
    status: Optional[str]      = Query(None),
    since: Optional[datetime]  = Query(None),
    until: Optional[datetime]  = Query(None),
    limit: int                 = Query(100, le=500),
    order: Optional[str]       = Query(None),
    before_id: Optional[int]   = Query(
        None,
        description="Cursor pagination: return alerts strictly older than "
                    "this alert id (keyset on created_at,id — no OFFSET). "
                    "Designed for order=recent; with severity ordering it "
                    "acts as an older-than filter."),
    lite: bool                 = Query(
        False,
        description="Lightweight rows only (id, created_at, status, "
                    "detection_type, severity_label, camera_id, store_id) — "
                    "no titles, snapshots, clip URLs, or extra payloads. "
                    "The default (false) payload is unchanged."),
):
    q = (db.query(Alert, DetectionEvent, Camera)
           .join(DetectionEvent, Alert.event_id == DetectionEvent.id)
           .outerjoin(Camera, DetectionEvent.camera_id == Camera.id))
    if before_id is not None:
        cursor = db.get(Alert, before_id)
        if cursor is not None:
            from sqlalchemy import tuple_
            q = q.filter(tuple_(Alert.created_at, Alert.id)
                         < tuple_(cursor.created_at, cursor.id))
    if status:
        q = q.filter(Alert.status == status)
    if camera_id:
        q = q.filter(DetectionEvent.camera_id == camera_id)
    if store_id is not None:
        q = q.filter(Camera.store_id == store_id)
    if detection_type:
        q = q.filter(DetectionEvent.detection_type == detection_type)
    if zone_id:
        q = q.filter(DetectionEvent.zone_id == zone_id)
    if since:
        q = q.filter(DetectionEvent.timestamp >= since)
    if until:
        q = q.filter(DetectionEvent.timestamp <= until)
    if order == "recent":
        # Pure recency — used by the real-time notification poller. Severity-
        # first ordering (below) would bury a NEW low-severity alert past
        # `limit` behind a backlog of higher-severity "new" alerts, so the
        # notifier never sees it and no sound fires in "All alerts" mode.
        q = q.order_by(desc(Alert.created_at)).limit(limit)
    else:
        # Order by severity rank (critical → warning → info) THEN recency, so a
        # flood of high-frequency info alerts (e.g. checkout_dwell) can't push
        # trespass / intrusion / shrinkage out of the returned window. When the
        # caller filters to one detection_type the rank is uniform and this
        # collapses to pure recency order.
        severity_rank = case(
            (DetectionEvent.detection_type.in_(_RANK_CRITICAL), 0),
            (DetectionEvent.detection_type.in_(_RANK_HIGH), 1),
            else_=2,
        )
        q = q.order_by(severity_rank.asc(), desc(Alert.created_at)).limit(limit)

    rows = q.all()

    # Lite mode — slim rows for pollers/badges that only need counts and
    # identity. Returned as a raw JSONResponse so the AlertOut
    # response_model never touches (or pads) it; the default path below
    # is byte-identical to before this flag existed.
    if lite:
        return JSONResponse([
            {
                "id":             a.id,
                "created_at":     (a.created_at.isoformat()
                                   if a.created_at else None),
                "status":         a.status,
                "detection_type": ev.detection_type,
                "severity_label": _severity_label(ev.detection_type, ev),
                "camera_id":      ev.camera_id,
                "store_id":       cam.store_id if cam else None,
            }
            for a, ev, cam in rows
        ])

    # Bulk-fetch zones AND stores referenced by these events so
    # _to_alert_out gets the names without N round-trips. Skipped
    # when no event in the page references one.
    zone_ids  = {ev.zone_id for _, ev, _ in rows if ev.zone_id is not None}
    store_ids = {cam.store_id for _, _, cam in rows if cam and cam.store_id is not None}
    zones_by_id  = ({z.id: z for z in db.query(Zone).filter(Zone.id.in_(zone_ids)).all()}
                    if zone_ids else {})
    from app.models import Store as _Store
    stores_by_id = ({s.id: s for s in db.query(_Store).filter(_Store.id.in_(store_ids)).all()}
                    if store_ids else {})

    return [
        _to_alert_out(
            a, ev, cam,
            zones_by_id.get(ev.zone_id) if ev.zone_id else None,
            stores_by_id.get(cam.store_id) if (cam and cam.store_id) else None,
        )
        for a, ev, cam in rows
    ]


@router.get("/{alert_id}", response_model=AlertOut)
def get_alert(alert_id: int, db: Session = Depends(get_db),
              _u: User = Depends(get_current_user)):
    """Single alert by id — same shape as the list endpoint, including
    clip_url and the snapshot fields. Registered AFTER the string routes
    (/summary, /export.xlsx) so those still match first. 404 when the alert
    or its underlying event is missing."""
    from app.models import Store as _Store
    a = db.get(Alert, alert_id)
    if not a:
        raise HTTPException(404, "alert not found")
    ev = db.get(DetectionEvent, a.event_id)
    if not ev:
        raise HTTPException(404, "alert event not found")
    cam = db.get(Camera, ev.camera_id) if ev.camera_id else None
    zone = db.get(Zone, ev.zone_id) if ev.zone_id else None
    store = db.get(_Store, cam.store_id) if (cam and cam.store_id) else None
    return _to_alert_out(a, ev, cam, zone, store)


@router.post("/{alert_id}/confirm", response_model=AlertActionOut)
def confirm(alert_id: int, db: Session = Depends(get_db),
            user: User = Depends(require_role("admin", "operator"))):
    # Single source of truth for confirm/dismiss/sprint-label.
    # See app/services/alert_feedback.py for the rationale.
    from app.services.alert_feedback import record_verdict
    return record_verdict(db, alert_id, "confirm", user)


@router.post("/{alert_id}/dismiss", response_model=AlertActionOut)
def dismiss(alert_id: int, db: Session = Depends(get_db),
            user: User = Depends(require_role("admin", "operator"))):
    from app.services.alert_feedback import record_verdict
    return record_verdict(db, alert_id, "dismiss", user)


@router.post("/{alert_id}/resolve", response_model=AlertActionOut)
def resolve(alert_id: int, db: Session = Depends(get_db),
            user: User = Depends(require_role("admin", "operator"))):
    """Mark an alert as resolved. Distinct from /confirm (which feeds
    training as a true positive); /resolve just records that the
    operator dealt with whatever was happening on the ground. Use
    this for routine operational alerts (queue, staff_present, …) so
    confirm/dismiss can stay narrowly about ML feedback quality."""
    a = db.get(Alert, alert_id)
    if not a:
        raise HTTPException(404, "alert not found")
    a.status = "confirmed"     # reuse the bucket — frontend treats as resolved
    a.assigned_to = user.id
    a.acknowledged_at = datetime.now(timezone.utc)
    a.resolved_at = datetime.now(timezone.utc)
    db.commit()
    return AlertActionOut(id=a.id, status=a.status)


@router.post("/resolve-all")
def resolve_all(db: Session = Depends(get_db),
                user: User = Depends(require_role("admin", "operator")),
                store_id: Optional[int] = Query(None),
                since: Optional[datetime] = Query(None),
                until: Optional[datetime] = Query(None)):
    """Bulk-mark every currently-NEW alert in the window (and optional
    store filter) as resolved. Mirrors the per-alert /resolve action.
    Returns the count actually flipped."""
    q = (db.query(Alert)
           .join(DetectionEvent, Alert.event_id == DetectionEvent.id)
           .outerjoin(Camera, DetectionEvent.camera_id == Camera.id)
           .filter(Alert.status == "new"))
    if store_id is not None:
        q = q.filter(Camera.store_id == store_id)
    if since:
        q = q.filter(DetectionEvent.timestamp >= since)
    if until:
        q = q.filter(DetectionEvent.timestamp <= until)
    now = datetime.now(timezone.utc)
    n = 0
    for a in q.all():
        a.status = "confirmed"   # same bucket the per-alert /resolve uses
        a.assigned_to = user.id
        a.acknowledged_at = now
        a.resolved_at = now
        n += 1
    db.commit()
    return {"resolved": n}


@router.post("/{alert_id}/note", response_model=AlertOut)
def add_note(alert_id: int, body: AlertNoteIn,
             db: Session = Depends(get_db),
             user: User = Depends(require_role("admin", "operator"))):
    """Append an investigation note. We APPEND rather than overwrite
    so multiple operators can leave a trail. Each entry is stamped
    with the username + timestamp so the trail reads chronologically."""
    a = db.get(Alert, alert_id)
    if not a:
        raise HTTPException(404, "alert not found")
    when = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    line = f"[{when}] {user.email}: {body.note.strip()}"
    a.notes = f"{a.notes}\n{line}" if a.notes else line
    db.commit()
    db.refresh(a)
    # Re-fetch with joins so the response shape matches /alerts.
    from app.models import Store as _Store
    ev = db.get(DetectionEvent, a.event_id)
    if not ev:
        # Orphaned alert — its DetectionEvent was pruned. _to_alert_out
        # dereferences event.camera_id unguarded, so bail with a clean
        # 404 instead of a 500.
        raise HTTPException(404, "alert event not found")
    cam = db.get(Camera, ev.camera_id) if ev.camera_id else None
    zone = db.get(Zone, ev.zone_id) if ev.zone_id else None
    store = db.get(_Store, cam.store_id) if (cam and cam.store_id) else None
    return _to_alert_out(a, ev, cam, zone, store)


@router.get("/{alert_id}/snapshot")
def alert_snapshot(alert_id: int, db: Session = Depends(get_db),
                   _u: User = Depends(get_current_user)):
    """Return a JPEG snapshot for the alert.

    Strategy:
      1. If the event has a stored thumbnail_path on disk, serve that
         (the moment-of-detection frame).
      2. Otherwise fall back to the camera's most recent cached frame
         from the streamer (vg:frame:{camera_id} in Redis). Stale by
         a few seconds but always available for active cameras.
      3. Otherwise 404 — frontend renders the placeholder.
    """
    from fastapi.responses import Response, FileResponse
    from app.stream.frame_buffer import FrameBuffer
    import os

    a = db.get(Alert, alert_id)
    if not a:
        raise HTTPException(404, "alert not found")
    ev = db.get(DetectionEvent, a.event_id)
    if not ev:
        raise HTTPException(404, "event not found")

    # Tier 1: archived thumbnail (if the inference worker stored one).
    if ev.thumbnail_path and os.path.exists(ev.thumbnail_path):
        return FileResponse(ev.thumbnail_path, media_type="image/jpeg")

    # Tier 2: camera's latest cached frame from the streamer.
    if ev.camera_id:
        cached = FrameBuffer().latest_jpeg(ev.camera_id)
        if cached:
            return Response(content=cached, media_type="image/jpeg")

    raise HTTPException(404, "no snapshot available")


@router.get("/{alert_id}/snapshot/{idx}")
def alert_snapshot_at(alert_id: int, idx: int,
                      db: Session = Depends(get_db),
                      _u: User = Depends(get_current_user)):
    """Return the Nth checkout-dwell timeline snapshot for the alert.
    Indices are 0-based and chronological (filenames embed the
    capture epoch). 404 on missing alert / out-of-range index / file
    pruned. Hidden filesystem path — operators only see indices."""
    from fastapi.responses import FileResponse
    import os
    a = db.get(Alert, alert_id)
    if not a:
        raise HTTPException(404, "alert not found")
    paths = list(a.snapshot_paths or [])
    if idx < 0 or idx >= len(paths):
        raise HTTPException(404, "snapshot index out of range")
    target = paths[idx]
    if not os.path.exists(target):
        raise HTTPException(404, "snapshot file missing (pruned?)")
    return FileResponse(target, media_type="image/jpeg")


@router.get("/{alert_id}/clip")
def alert_clip(alert_id: int, db: Session = Depends(get_db),
               _u: User = Depends(get_current_user)):
    """Stream the recorded video clip for an alert. The path lives in the
    event's extra JSON (written by recorder.extract_pending_clips). 404 when
    the alert has no clip or it was pruned (48h retention). Auth required."""
    from fastapi.responses import FileResponse
    import os
    a = db.get(Alert, alert_id)
    if not a:
        raise HTTPException(404, "alert not found")
    ev = db.get(DetectionEvent, a.event_id)
    path = (ev.extra or {}).get("alert_clip_path") if ev else None
    if not path or not os.path.exists(path):
        raise HTTPException(404, "clip not available")
    return FileResponse(path, media_type="video/mp4",
                        filename=f"alert_{alert_id}.mp4")
